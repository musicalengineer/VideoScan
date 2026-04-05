#!/usr/bin/env python3
"""
scan_videos.py — Broad Video File Cataloger
============================================
Scans a volume or directory for all video files and produces a detailed
Excel spreadsheet with codec, resolution, duration, dates, and more.

Uses ffprobe (bundled with ffmpeg) for all metadata — no other dependencies
except openpyxl for the spreadsheet output.

Requirements:
    brew install ffmpeg
    pip install openpyxl

Usage:
    python scan_videos.py /Volumes/Seagate2TB
    python scan_videos.py /Volumes/Seagate2TB --output my_catalog.xlsx
    python scan_videos.py ~/Movies
"""

import os
import sys
import json
import argparse
import hashlib
import datetime
import subprocess
from pathlib import Path
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not found.  Run:  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Video extensions to scan
# ---------------------------------------------------------------------------
VIDEO_EXTENSIONS = {
    ".mov", ".mp4", ".m4v", ".avi", ".mkv", ".mxf",
    ".mts", ".m2ts", ".ts", ".mpg", ".mpeg", ".m2v", ".vob",
    ".wmv", ".asf", ".webm", ".ogv", ".ogg",
    ".rm", ".rmvb", ".divx", ".flv", ".f4v",
    ".3gp", ".3g2", ".dv", ".dif",
    ".braw", ".r3d",
    ".vro", ".mod", ".tod",
}

SKIP_DIRS = {
    ".spotlight-v100", ".fseventsd", ".trashes", ".temporaryitems",
    ".documentrevisions-v100", ".vol", "automount",
}

# ---------------------------------------------------------------------------
# ffprobe
# ---------------------------------------------------------------------------

def check_ffprobe():
    try:
        r = subprocess.run(["ffprobe", "-version"], capture_output=True)
        return r.returncode == 0
    except FileNotFoundError:
        return False


def run_ffprobe(filepath):
    """Run ffprobe on a file, return parsed JSON or None."""
    cmd = [
        "ffprobe",
        "-v", "quiet",
        "-print_format", "json",
        "-show_format",
        "-show_streams",
        str(filepath),
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            return None
        return json.loads(result.stdout)
    except (subprocess.TimeoutExpired, json.JSONDecodeError, Exception):
        return None


def extract_metadata(probe):
    """Pull the fields we care about out of ffprobe JSON."""
    meta = {
        "duration":        "",
        "container":       "",
        "video_codec":     "",
        "width":           "",
        "height":          "",
        "resolution":      "",
        "frame_rate":      "",
        "video_bitrate":   "",
        "total_bitrate":   "",
        "color_space":     "",
        "bit_depth":       "",
        "scan_type":       "",
        "audio_codec":     "",
        "audio_channels":  "",
        "audio_samplerate":"",
        "timecode":        "",
        "tape_name":       "",
        "creation_time":   "",
        "is_playable":     "Yes",   # if ffprobe read it, it's playable
        "stream_type":     "",
        "notes":           "",
    }

    fmt     = probe.get("format", {})
    streams = probe.get("streams", [])
    tags    = fmt.get("tags", {}) or {}

    # Container / format
    meta["container"] = fmt.get("format_long_name") or fmt.get("format_name", "")

    # Duration
    dur = fmt.get("duration")
    if dur:
        try:
            secs = float(dur)
            meta["duration"] = format_duration(secs)
        except ValueError:
            pass

    # Total bitrate
    br = fmt.get("bit_rate")
    if br:
        try:
            meta["total_bitrate"] = f"{int(br) // 1000} kbps"
        except ValueError:
            pass

    # Format-level tags
    meta["timecode"]      = (tags.get("timecode") or tags.get("Timecode") or "").strip()
    meta["tape_name"]     = (tags.get("tape_name") or tags.get("reel_name") or
                             tags.get("com.apple.quicktime.reelname") or "").strip()
    meta["creation_time"] = (tags.get("creation_time") or
                             tags.get("com.apple.quicktime.creationdate") or "").strip()

    has_video = False
    has_audio = False

    for s in streams:
        ctype = s.get("codec_type", "")
        stags = s.get("tags", {}) or {}

        if not meta["timecode"]:
            meta["timecode"] = (stags.get("timecode") or "").strip()
        if not meta["creation_time"]:
            meta["creation_time"] = (stags.get("creation_time") or "").strip()

        if ctype == "video" and not has_video:
            has_video = True
            meta["video_codec"] = s.get("codec_name", "")
            meta["width"]       = str(s.get("width", ""))
            meta["height"]      = str(s.get("height", ""))
            if meta["width"] and meta["height"]:
                meta["resolution"] = f"{meta['width']}x{meta['height']}"

            # Frame rate — stored as fraction e.g. "30000/1001"
            fr = s.get("r_frame_rate", "") or s.get("avg_frame_rate", "")
            if "/" in fr:
                try:
                    n, d = fr.split("/")
                    fps = float(n) / float(d)
                    meta["frame_rate"] = f"{fps:.3f}".rstrip("0").rstrip(".")
                except Exception:
                    meta["frame_rate"] = fr
            else:
                meta["frame_rate"] = fr

            vbr = s.get("bit_rate")
            if vbr:
                try:
                    meta["video_bitrate"] = f"{int(vbr) // 1000} kbps"
                except ValueError:
                    pass

            meta["color_space"] = s.get("color_space", "")
            meta["bit_depth"]   = str(s.get("bits_per_raw_sample", ""))
            meta["scan_type"]   = s.get("field_order", "")

        elif ctype == "audio" and not has_audio:
            has_audio = True
            meta["audio_codec"]      = s.get("codec_name", "")
            meta["audio_channels"]   = str(s.get("channels", ""))
            sr = s.get("sample_rate")
            if sr:
                meta["audio_samplerate"] = f"{sr} Hz"

    # Stream type classification (especially useful for MXF)
    if has_video and has_audio:
        meta["stream_type"] = "Video+Audio"
    elif has_video:
        meta["stream_type"] = "Video only"
    elif has_audio:
        meta["stream_type"] = "Audio only"
    else:
        meta["stream_type"] = "No A/V streams"
        meta["is_playable"] = "No streams found"

    return meta

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def human_size(nb):
    for u in ("B", "KB", "MB", "GB", "TB"):
        if abs(nb) < 1024.0:
            return f"{nb:.1f} {u}"
        nb /= 1024.0
    return f"{nb:.1f} PB"


def format_duration(secs):
    secs = int(secs)
    h, rem = divmod(secs, 3600)
    m, s   = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def partial_md5(filepath, chunk=65536):
    h = hashlib.md5()
    try:
        size = os.path.getsize(filepath)
        with open(filepath, "rb") as f:
            h.update(f.read(chunk))
            if size > chunk * 2:
                f.seek(-chunk, 2)
                h.update(f.read(chunk))
        return h.hexdigest()
    except Exception:
        return ""

# ---------------------------------------------------------------------------
# Scan
# ---------------------------------------------------------------------------

def scan(root):
    root = Path(root).resolve()
    found = []
    skipped = []

    print(f"\nScanning: {root}")
    for dirpath, dirnames, filenames in os.walk(root, followlinks=False):
        dirnames[:] = [
            d for d in dirnames
            if d.lower() not in SKIP_DIRS and not d.startswith(".")
        ]
        for fname in filenames:
            if Path(fname).suffix.lower() in VIDEO_EXTENSIONS:
                found.append(Path(dirpath) / fname)

    total = len(found)
    print(f"Found {total} video files. Probing with ffprobe...\n")

    records = []
    for i, fp in enumerate(found, 1):
        print(f"  [{i}/{total}] {fp.name[:60]:<60}", end="\r")
        try:
            stat     = fp.stat()
            size     = stat.st_size
            mtime    = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            ctime_ts = stat.st_birthtime if hasattr(stat, "st_birthtime") else stat.st_ctime
            ctime    = datetime.datetime.fromtimestamp(ctime_ts).strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            skipped.append(f"{fp} — stat failed: {e}")
            continue

        probe = run_ffprobe(fp)
        if probe is None:
            meta = {k: "" for k in [
                "duration","container","video_codec","width","height","resolution",
                "frame_rate","video_bitrate","total_bitrate","color_space","bit_depth",
                "scan_type","audio_codec","audio_channels","audio_samplerate",
                "timecode","tape_name","creation_time","stream_type","notes",
            ]}
            meta["is_playable"] = "ffprobe failed"
            meta["notes"]       = "ffprobe could not read file"
        else:
            meta = extract_metadata(probe)

        records.append({
            "Filename":          fp.name,
            "Extension":         fp.suffix.upper().lstrip("."),
            "Stream Type":       meta["stream_type"],
            "Size":              human_size(size),
            "Size (Bytes)":      size,
            "Duration":          meta["duration"],
            "Date Created":      ctime,
            "Date Modified":     mtime,
            "Container":         meta["container"],
            "Video Codec":       meta["video_codec"],
            "Resolution":        meta["resolution"],
            "Frame Rate":        meta["frame_rate"],
            "Video Bitrate":     meta["video_bitrate"],
            "Total Bitrate":     meta["total_bitrate"],
            "Color Space":       meta["color_space"],
            "Bit Depth":         meta["bit_depth"],
            "Scan Type":         meta["scan_type"],
            "Audio Codec":       meta["audio_codec"],
            "Audio Channels":    meta["audio_channels"],
            "Audio Sample Rate": meta["audio_samplerate"],
            "Timecode":          meta["timecode"],
            "Tape Name":         meta["tape_name"],
            "Is Playable":       meta["is_playable"],
            "Partial MD5":       partial_md5(str(fp)),
            "Full Path":         str(fp),
            "Directory":         str(fp.parent),
            "Notes":             meta["notes"],
        })

    print()
    return records, skipped

# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Filename",          32), ("Extension",         8), ("Stream Type",      13),
    ("Size",              11), ("Size (Bytes)",      14), ("Duration",         10),
    ("Date Created",      18), ("Date Modified",     18), ("Container",        22),
    ("Video Codec",       13), ("Resolution",        11), ("Frame Rate",       10),
    ("Video Bitrate",     13), ("Total Bitrate",     13), ("Color Space",      12),
    ("Bit Depth",          9), ("Scan Type",         11), ("Audio Codec",      11),
    ("Audio Channels",    13), ("Audio Sample Rate", 15), ("Timecode",         13),
    ("Tape Name",         18), ("Is Playable",       12), ("Partial MD5",      34),
    ("Full Path",         60), ("Directory",         50), ("Notes",            25),
]

# Color by stream type
STREAM_COLORS = {
    "Video+Audio":       "DDEEFF",
    "Video only":        "FFF0CC",
    "Audio only":        "FFE0E0",
    "No A/V streams":    "EEEEEE",
    "ffprobe failed":    "FFD0D0",
}

HDR_BLUE  = "1F3864"
ALT_GREY  = "F5F8FF"


def write_xlsx(records, skipped, root, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Video Catalog"

    thin  = Side(style="thin", color="CCCCCC")
    bd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color=HDR_BLUE, end_color=HDR_BLUE)
    halign= Alignment(horizontal="center", vertical="center", wrap_text=True)
    dfont = Font(name="Arial", size=9)

    # Title
    ncols = len(COLUMNS)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value     = (f"Video Catalog  —  {root}  —  "
                   f"{len(records)} files  —  "
                   f"Scanned {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    t.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="0D1F2D", end_color="0D1F2D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for ci, (label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = bd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # Data
    LEFT  = {"Filename", "Full Path", "Directory", "Container", "Notes", "Tape Name"}
    for ri, rec in enumerate(records, 3):
        st = rec.get("Stream Type", "")
        fc = STREAM_COLORS.get(st, "FFFFFF")
        if ri % 2 == 0 and fc == "FFFFFF":
            fc = ALT_GREY
        fill = PatternFill("solid", start_color=fc, end_color=fc)

        for ci, (col, _) in enumerate(COLUMNS, 1):
            val  = rec.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = dfont
            cell.fill   = fill
            cell.border = bd
            cell.alignment = Alignment(
                horizontal="left" if col in LEFT else "center",
                vertical="top"
            )

    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{len(records)+2}"

    # --- Summary sheet ---
    ss = wb.create_sheet("Summary")
    ss.column_dimensions["A"].width = 28
    ss.column_dimensions["B"].width = 18
    ss.column_dimensions["C"].width = 40

    ext_counts    = Counter(r["Extension"] for r in records)
    stream_counts = Counter(r["Stream Type"] for r in records)
    codec_counts  = Counter(r["Video Codec"] for r in records if r["Video Codec"])
    total_bytes   = sum(r["Size (Bytes)"] for r in records)
    playable      = sum(1 for r in records if r["Is Playable"] == "Yes")
    ffprobe_fail  = sum(1 for r in records if "ffprobe" in r["Is Playable"])

    def sh(row, col, val, bold=False, color=None):
        c = ss.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=bold, size=10,
                      color=color if color else "000000")
        c.alignment = Alignment(horizontal="left", vertical="center")
        return c

    def section(row, title):
        c = ss.cell(row=row, column=1, value=title)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=HDR_BLUE, end_color=HDR_BLUE)
        ss.merge_cells(f"A{row}:C{row}")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ss.row_dimensions[row].height = 20

    r = 1
    section(r, "SCAN SUMMARY"); r+=1
    sh(r,1,"Volume Scanned",bold=True); sh(r,2,str(root)); r+=1
    sh(r,1,"Scan Date",bold=True);      sh(r,2,datetime.datetime.now().strftime("%Y-%m-%d %H:%M")); r+=1
    sh(r,1,"Total Files Found",bold=True); sh(r,2,len(records)); r+=1
    sh(r,1,"Total Size",bold=True);     sh(r,2,human_size(total_bytes)); r+=1
    sh(r,1,"Playable (ffprobe OK)",bold=True); sh(r,2,playable); r+=1
    sh(r,1,"ffprobe Could Not Read",bold=True); sh(r,2,ffprobe_fail,color="AA0000" if ffprobe_fail else "000000"); r+=1

    r+=1
    section(r, "BY STREAM TYPE"); r+=1
    for stype, cnt in stream_counts.most_common():
        sh(r,1,stype,bold=True); sh(r,2,cnt); r+=1

    r+=1
    section(r, "BY EXTENSION"); r+=1
    for ext, cnt in ext_counts.most_common():
        sh(r,1,ext); sh(r,2,cnt); r+=1

    r+=1
    section(r, "BY VIDEO CODEC"); r+=1
    for codec, cnt in codec_counts.most_common():
        sh(r,1,codec); sh(r,2,cnt); r+=1

    # Duplicate hint
    from collections import Counter as C2
    hash_counts = C2(r["Partial MD5"] for r in records if r["Partial MD5"])
    dups = {h: c for h, c in hash_counts.items() if c > 1}
    if dups:
        r+=1
        section(r, "POTENTIAL DUPLICATES"); r+=1
        dup_files = sum(dups.values())
        sh(r,1,"Files sharing a partial MD5",bold=True)
        sh(r,2,dup_files)
        sh(r,3,"Filter 'Partial MD5' column in catalog to find them")
        r+=1

    # Skipped files
    if skipped:
        sk = wb.create_sheet("Skipped")
        sk["A1"].value = "Files skipped due to permission or read errors"
        sk["A1"].font  = Font(name="Arial", bold=True)
        for i, path in enumerate(skipped, 2):
            sk.cell(row=i, column=1, value=path).font = Font(name="Arial", size=9)
        sk.column_dimensions["A"].width = 80

    wb.save(output_path)
    print(f"Spreadsheet saved → {output_path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Scan a volume for video files and export a detailed Excel catalog.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scan_videos.py /Volumes/Seagate2TB
  python scan_videos.py ~/Movies --output movies.xlsx
  python scan_videos.py /Volumes/Seagate2TB --no-hash
        """,
    )
    parser.add_argument("volume",    help="Path to scan")
    parser.add_argument("--output",  "-o", default="", help="Output .xlsx filename")
    parser.add_argument("--no-hash", action="store_true",
                        help="Skip partial MD5 hashing (faster, loses duplicate detection)")
    args = parser.parse_args()

    if not os.path.isdir(args.volume):
        print(f"ERROR: '{args.volume}' is not a valid directory.")
        sys.exit(1)

    if not check_ffprobe():
        print("ERROR: ffprobe not found.  Install with:  brew install ffmpeg")
        sys.exit(1)

    output = args.output
    if not output:
        safe  = args.volume.replace("/", "_").replace(" ", "_").strip("_") or "root"
        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output = f"video_catalog_{safe}_{ts}.xlsx"

    records, skipped = scan(args.volume)

    if not records:
        print("No video files found.")
        sys.exit(0)

    if args.no_hash:
        for r in records:
            r["Partial MD5"] = ""

    print(f"Writing spreadsheet for {len(records)} files...")
    write_xlsx(records, skipped, args.volume, output)

    # Console summary
    sc = Counter(r["Stream Type"] for r in records)
    print(f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Scan Complete
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Total files:    {len(records)}
  Video+Audio:    {sc.get('Video+Audio', 0)}
  Video only:     {sc.get('Video only', 0)}
  Audio only:     {sc.get('Audio only', 0)}
  ffprobe failed: {sum(1 for r in records if 'ffprobe' in r['Is Playable'])}
  Output:         {output}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
""")

if __name__ == "__main__":
    main()